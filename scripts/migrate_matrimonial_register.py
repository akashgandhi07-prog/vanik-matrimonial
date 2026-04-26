#!/usr/bin/env python3
"""
Import members from Matrimonial_Register.xlsx into Supabase (profiles + member_private + auth).

Requires Python 3.10+. Install: pip install -r scripts/requirements-migration.txt

Env (see .env.example):
  SUPABASE_URL (or VITE_SUPABASE_URL)
  SUPABASE_SERVICE_ROLE_KEY
  Optional welcome email: RESEND_API_KEY, PUBLIC_SITE_URL

Usage:
  python scripts/migrate_matrimonial_register.py --dry-run
  python scripts/migrate_matrimonial_register.py --ref "M 2516"
  python scripts/migrate_matrimonial_register.py --excel /path/to/Matrimonial_Register.xlsx
  python scripts/migrate_matrimonial_register.py --send-welcome-emails
  python scripts/migrate_matrimonial_register.py --inspect-headers
"""

from __future__ import annotations

import argparse
import html
import json
import logging
import math
import re
import secrets
import string
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from typing import Any, Iterable

from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None  # type: ignore[misc, assignment]

from supabase import Client, create_client

# ---------------------------------------------------------------------------
# Column layout (1-based Excel letters). Boys vs Girls differ for mobile / home.
# ---------------------------------------------------------------------------

BOYS_COLS = {
    "reference": "B",
    "age_check": "C",
    "education": "D",
    "job_title": "E",
    "height": "F",
    "hobbies": "G",
    "diet": "H",
    "religion": "I",
    "nationality": "J",
    "order_date": "K",
    "first_name": "O",
    "surname": "P",
    "email": "Q",
    "mobile": "AD",
    "gender": "AE",
    "community": "AN",
    "future_settlement": "AP",
    "town_country_of_origin": "AQ",
    "place_of_birth": "AR",
    "dob": "AV",
    "father": "AW",
    "mother": "AX",
    "show": "AZ",
}

def girls_columns(mobile_col: str = "AC") -> dict[str, str]:
    """Girls sheet: AD is home address; mobile column differs from Boys (AD). Override via --girls-mobile-col."""
    return {
        **{k: v for k, v in BOYS_COLS.items() if k != "mobile"},
        "mobile": mobile_col,
        "home_address": "AD",
    }

RESEND_FROM = "Vanik Matrimonial Register <noreply@vanikmatrimonial.co.uk>"

MONTHS = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def col_idx(letter: str) -> int:
    from openpyxl.utils.cell import column_index_from_string

    return column_index_from_string(letter.upper())


def cell(ws: Any, row: int, letter: str) -> Any:
    return ws.cell(row=row, column=col_idx(letter)).value


def sheet_by_name(wb: Any, *names: str) -> Any:
    lowered = {s.strip().lower(): s for s in wb.sheetnames}
    for n in names:
        key = n.strip().lower()
        if key in lowered:
            return wb[lowered[key]]
    raise KeyError(f"No sheet matching {names!r}; have {wb.sheetnames!r}")


def str_clean(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, float) and not math.isnan(v):
        if v == int(v):
            return str(int(v))
        return str(v).strip()
    s = str(v).strip()
    return s or None


def excel_bool_true(v: Any) -> bool:
    if v is True:
        return True
    if v is False or v is None:
        return False
    s = str_clean(v)
    if not s:
        return False
    return s.upper() in ("TRUE", "YES", "1", "Y")


def normalise_reference(ref: Any, sheet_gender: str) -> str | None:
    s = str_clean(ref)
    if not s:
        return None
    s = s.upper().replace(" ", "")
    m = re.match(r"^([MF])(\d+)$", s)
    if m:
        return f"{m.group(1)} {m.group(2)}"
    m = re.match(r"^([MF])\s+(\d+)$", s.upper())
    if m:
        return f"{m.group(1)} {m.group(2)}"
    prefix = "M" if sheet_gender.lower() == "male" else "F"
    if re.match(r"^\d+$", s):
        return f"{prefix} {s}"
    return None


def normalise_gender(raw: Any, fallback: str) -> str | None:
    s = str_clean(raw)
    if not s:
        return fallback if fallback in ("Male", "Female") else None
    sl = s.lower()
    if sl.startswith("m"):
        return "Male"
    if sl.startswith("f"):
        return "Female"
    return fallback if fallback in ("Male", "Female") else None


def normalise_community(raw: Any) -> str:
    s = (str_clean(raw) or "").strip()
    if not s:
        return "Other"
    key = s.lower()
    if "vanik" in key:
        return "Vanik"
    if "lohana" in key:
        return "Lohana"
    if "brahmin" in key:
        return "Brahmin"
    if key == "other":
        return "Other"
    return "Other"


def normalise_religion(raw: Any) -> str:
    s = (str_clean(raw) or "").lower()
    if "jain" in s:
        return "Jain"
    if "hindu" in s:
        return "Hindu"
    return "Other"


def normalise_diet(raw: Any, ref: str, log: logging.Logger) -> str | None:
    s = (str_clean(raw) or "").strip()
    if not s:
        return None
    key = s.lower().replace(" ", "").replace("-", "")
    if key in ("veg", "vegetarian"):
        return "Veg"
    if key in ("nonveg", "nonveg."):
        return "Non-veg"
    if key == "vegan":
        return "Vegan"
    if key == "other":
        log.warning("Diet 'Other' mapped to Veg (uncertain) for ref %s", ref)
        return "Veg"
    if "non" in key and "veg" in key:
        return "Non-veg"
    log.warning("Unknown diet %r for ref %s - left unset", raw, ref)
    return None


def inches_to_cm(feet: int, inches: int) -> int:
    total = feet * 12 + min(max(inches, 0), 11)
    return int(round(total * 2.54))


def parse_height_cm(raw: Any) -> tuple[int | None, str | None]:
    """Return (cm, original_string_if_failed)."""
    if raw is None:
        return None, None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        n = float(raw)
        if 120 <= n <= 230:
            return int(round(n)), None
        if n < 10:  # unlikely feet as float alone
            return inches_to_cm(int(n), 0), None
        return None, str(raw)
    s0 = str(raw).strip()
    if not s0:
        return None, None
    s = s0.replace("″", '"').replace("′", "'").replace("’", "'").replace("“", '"')

    # Already centimetres (plain integer)
    m = re.match(r"^(\d{2,3})\s*cm\s*$", s, re.I)
    if m:
        v = int(m.group(1))
        if 120 <= v <= 230:
            return v, None

    m = re.match(r"^(\d{2,3})$", s)
    if m:
        v = int(m.group(1))
        if 120 <= v <= 230:
            return v, None

    # 5ft 7, 5 ft 7
    m = re.match(r"(?i)^\s*(\d+)\s*ft\s*(\d{1,2})?\s*$", s)
    if m:
        ft = int(m.group(1))
        inch = int(m.group(2) or 0)
        return inches_to_cm(ft, inch), None

    # 5'10, 5' 9, 5'10", 6'
    m = re.match(
        r"(?i)^\s*(\d+)\s*['\u2032\u2019]\s*(\d{0,2})\s*([\"\u2033\u201d]?)\s*$",
        s,
    )
    if m:
        ft = int(m.group(1))
        inch_s = m.group(2) or ""
        inch = int(inch_s) if inch_s else 0
        return inches_to_cm(ft, inch), None

    # 5,11 or 5.11 as feet, inches (second part 0-11)
    m = re.match(r"(?i)^\s*(\d+)\s*[,]\s*(\d{1,2})\s*$", s)
    if m:
        ft, inch = int(m.group(1)), int(m.group(2))
        if inch <= 11:
            return inches_to_cm(ft, inch), None

    m = re.match(r"(?i)^\s*(\d+)\s*\.\s*(\d{1,2})\s*$", s)
    if m:
        ft, inch = int(m.group(1)), int(m.group(2))
        if inch <= 11:
            return inches_to_cm(ft, inch), None

    return None, s0


def two_digit_year(yy: int) -> int:
    """Heuristic: matrimonial ages → 1930s-2010s."""
    if yy >= 30:
        return 1900 + yy
    return 2000 + yy


def _from_excel_serial(raw: Any) -> date | None:
    if not isinstance(raw, (int, float)) or isinstance(raw, bool):
        return None
    n = float(raw)
    if not (200 < n < 1_200_000):
        return None
    try:
        from openpyxl.utils.datetime import from_excel

        dt = from_excel(n)
        if isinstance(dt, datetime):
            return dt.date()
        if isinstance(dt, date):
            return dt
    except Exception:
        return None
    return None


def parse_date_of_birth(raw: Any, ref: str, log: logging.Logger) -> date | None:
    if raw is None:
        log.error("DOB empty for ref %s", ref)
        return None
    ser = _from_excel_serial(raw)
    if ser is not None:
        return ser
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str_clean(raw)
    if not s:
        log.error("DOB empty for ref %s", ref)
        return None

    # Nov-99, Jul-1999
    m = re.match(
        r"(?i)^\s*([a-z]{3,9})\s*[-/]\s*(\d{2,4})\s*$",
        s,
    )
    if m:
        mon_s, y_s = m.group(1).lower(), m.group(2)
        mon = MONTHS.get(mon_s[:3])
        if not mon:
            log.error("DOB unknown month %r for ref %s", raw, ref)
            return None
        if len(y_s) == 2:
            year = two_digit_year(int(y_s))
        else:
            year = int(y_s)
        return date(year, mon, 1)

    # 99-Nov style
    m = re.match(r"(?i)^\s*(\d{2,4})\s*[-/]\s*([a-z]{3,9})\s*$", s)
    if m:
        y_s, mon_s = m.group(1), m.group(2).lower()
        mon = MONTHS.get(mon_s[:3])
        if not mon:
            return None
        year = two_digit_year(int(y_s)) if len(y_s) == 2 else int(y_s)
        return date(year, mon, 1)

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    log.error("DOB unparseable %r for ref %s", raw, ref)
    return None


def parse_order_date(raw: Any) -> date | None:
    if raw is None:
        return None
    ser = _from_excel_serial(raw)
    if ser is not None:
        return ser
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str_clean(raw)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def membership_expires_at(order: date | None) -> datetime | None:
    if not order:
        return None
    end = order + timedelta(days=365)
    return datetime(end.year, end.month, end.day, 23, 59, 59, tzinfo=timezone.utc)


def random_password_16() -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(16))


def send_resend_email(api_key: str, to: str, subject: str, html_body: str) -> tuple[str | None, str | None]:
    payload = {
        "from": RESEND_FROM,
        "to": [to],
        "subject": subject,
        "html": html_body,
    }
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = json.loads(resp.read().decode())
            return body.get("id"), None
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode()
        except Exception:
            err_body = str(e)
        return None, err_body or str(e)


def welcome_email_html(first_name: str, reference: str, site_url: str, reset_link: str) -> str:
    fn = html.escape(first_name)
    ref = html.escape(reference)
    site = html.escape(site_url.rstrip("/"))
    link = html.escape(reset_link, quote=True)
    inner = f"""<p>Dear {fn},</p>
<p>We have moved the Vanik Matrimonial Register to a new online system. Your reference number is
<strong>{ref}</strong>.</p>
<p>You can sign in at <a href="{site}">{site}</a>. Please use the link below to choose your own password.
This link is generated by our auth provider; set your password promptly (we recommend within 24 hours).</p>
<p style="margin:24px 0;"><a href="{link}" style="display:inline-block;padding:12px 20px;background:#4f46e5;
color:#fff;text-decoration:none;border-radius:8px;font-weight:600;">Set your password</a></p>
<p>If the button does not work, copy and paste this URL into your browser:<br/><span style="word-break:break-all;
font-size:13px;">{reset_link}</span></p>
<p>Your profile details have been transferred from the previous register. If anything looks wrong, reply to this email.</p>
<p>With good wishes,<br/>The register team<br/>
<a href="mailto:mahesh.gandhi@vanikcouncil.uk">mahesh.gandhi@vanikcouncil.uk</a></p>"""
    return (
        "<!DOCTYPE html><html><head><meta charset=\"utf-8\"/></head><body style=\"margin:0;padding:24px;"
        "background:#f9fafb;font-family:system-ui,sans-serif;font-size:15px;line-height:1.6;color:#111827;\">"
        f"<div style=\"max-width:560px;margin:0 auto;background:#fff;border:1px solid #e5e7eb;"
        f"border-radius:12px;padding:28px 32px;\">{inner}"
        '<p style="margin:24px 0 0;font-size:13px;color:#6b7280;">Vanik Council UK - Vanik Matrimonial Register</p>'
        "</div></body></html>"
    )


@dataclass
class MigrationReport:
    imported_male: int = 0
    imported_female: int = 0
    membership_active: int = 0
    membership_expired: int = 0
    skipped_existing: int = 0
    skipped_not_show: int = 0
    skipped_missing: int = 0
    height_fail_refs: list[str] = field(default_factory=list)
    dob_fail_refs: list[str] = field(default_factory=list)
    diet_uncertain_refs: list[str] = field(default_factory=list)
    missing_photo_names: list[str] = field(default_factory=list)
    reset_links: list[tuple[str, str, str, str]] = field(
        default_factory=list
    )  # ref, display_name, email, link
    errors: list[str] = field(default_factory=list)

    def print_summary(self) -> None:
        total = self.imported_male + self.imported_female
        print("\n========== MIGRATION REPORT ==========")
        print(f"Total imported: {total} (M {self.imported_male}, F {self.imported_female})")
        print(f"Membership active: {self.membership_active}")
        print(f"Membership expired: {self.membership_expired}")
        print(f"Skipped (already in DB): {self.skipped_existing}")
        print(f"Skipped (Show? not TRUE): {self.skipped_not_show}")
        print(f"Skipped (missing required fields / errors): {self.skipped_missing}")
        if self.height_fail_refs:
            print(f"Height normalisation failures ({len(self.height_fail_refs)}): {', '.join(self.height_fail_refs)}")
        if self.dob_fail_refs:
            print(f"DOB parse failures ({len(self.dob_fail_refs)}): {', '.join(self.dob_fail_refs)}")
        if self.diet_uncertain_refs:
            print(f"Diet 'Other' → Veg ({len(self.diet_uncertain_refs)}): {', '.join(self.diet_uncertain_refs)}")
        if self.missing_photo_names:
            print(f"Missing photos (follow-up, n={len(self.missing_photo_names)}):")
            for n in self.missing_photo_names:
                print(f"  - {n}")
        print("\n--- Members needing password reset (recovery links) ---")
        for ref, name, email, link in self.reset_links:
            print(f"{ref} | {name} | {email}\n  {link}\n")
        if self.errors:
            print("\n--- Errors ---")
            for e in self.errors:
                print(f"  {e}")
        print(
            "Note: recovery links use your Supabase Auth mailer settings; set a short OTP expiry (e.g. 24h) in "
            "Dashboard → Authentication → Email → OTP expiry."
        )
        print("======================================\n")


def profile_exists(supabase: Client, reference_number: str) -> bool:
    r = (
        supabase.table("profiles")
        .select("id")
        .eq("reference_number", reference_number)
        .limit(1)
        .execute()
    )
    return bool(r.data)


def inspect_headers(path: str) -> None:
    from openpyxl.utils.cell import get_column_letter

    wb = load_workbook(path, read_only=True, data_only=True)
    for title in ("Boys", "Girls"):
        try:
            ws = sheet_by_name(wb, title)
        except KeyError as e:
            print(e)
            continue
        print(f"\n=== {title} sheet: row 1 headers (A-AZ) ===")
        for i in range(1, 52):
            letter = get_column_letter(i)
            val = ws.cell(row=1, column=i).value
            if val is not None and str(val).strip():
                print(f"  {letter}: {val!r}")


def process_sheet(
    ws: Any,
    cols: dict[str, str],
    sheet_gender: str,
    args: argparse.Namespace,
    supabase: Client | None,
    report: MigrationReport,
    log: logging.Logger,
) -> None:
    today = date.today()
    site_url = (args.public_site_url or "https://vanikmatrimonial.co.uk").rstrip("/")
    redirect_to = f"{site_url}/reset-password"

    max_row = ws.max_row or 0
    for row in range(2, max_row + 1):
        ref_raw = cell(ws, row, cols["reference"])
        ref = normalise_reference(ref_raw, sheet_gender)
        if not ref:
            continue
        if args.ref_filter:
            if ref.replace(" ", "").upper() != args.ref_filter.replace(" ", "").upper():
                continue

        if not excel_bool_true(cell(ws, row, cols["show"])):
            report.skipped_not_show += 1
            continue

        if not args.dry_run and supabase and profile_exists(supabase, ref):
            report.skipped_existing += 1
            log.info("Skip existing ref %s", ref)
            continue
        if args.dry_run:
            # optional read for duplicate message
            if supabase and profile_exists(supabase, ref):
                report.skipped_existing += 1
                log.info("[dry-run] Would skip existing ref %s", ref)
                continue

        first = str_clean(cell(ws, row, cols["first_name"]))
        surname = str_clean(cell(ws, row, cols["surname"]))
        email = str_clean(cell(ws, row, cols["email"]))
        mobile = str_clean(cell(ws, row, cols["mobile"]))
        dob_raw = cell(ws, row, cols["dob"])
        dob = parse_date_of_birth(dob_raw, ref, log)
        if dob is None:
            report.dob_fail_refs.append(ref)
            report.skipped_missing += 1
            continue
        if not first or not surname or not email or not mobile:
            report.skipped_missing += 1
            report.errors.append(f"{ref}: missing first/surname/email/mobile")
            continue
        if "@" not in email or "." not in email.split("@")[-1]:
            report.skipped_missing += 1
            report.errors.append(f"{ref}: invalid email {email!r}")
            continue

        order_d = parse_order_date(cell(ws, row, cols["order_date"]))
        mem_exp = membership_expires_at(order_d)
        is_expired = mem_exp is None or mem_exp.date() < today

        status = "expired" if is_expired else "active"
        show_reg = False if is_expired else True

        gender = normalise_gender(cell(ws, row, cols["gender"]), sheet_gender)
        if gender not in ("Male", "Female"):
            report.skipped_missing += 1
            report.errors.append(f"{ref}: invalid gender")
            continue

        height_cm, height_orig_fail = parse_height_cm(cell(ws, row, cols["height"]))
        if height_orig_fail:
            report.height_fail_refs.append(ref)
            log.warning("Height unrecognised %r for ref %s - storing null", height_orig_fail, ref)

        diet_raw = cell(ws, row, cols["diet"])
        diet = normalise_diet(diet_raw, ref, log)
        if str_clean(diet_raw) and str_clean(diet_raw).strip().lower() == "other":
            report.diet_uncertain_refs.append(ref)

        profile_row = {
            "reference_number": ref,
            "gender": gender,
            "first_name": first,
            "education": str_clean(cell(ws, row, cols["education"])),
            "job_title": str_clean(cell(ws, row, cols["job_title"])),
            "height_cm": height_cm,
            "hobbies": str_clean(cell(ws, row, cols["hobbies"])),
            "diet": diet,
            "religion": normalise_religion(cell(ws, row, cols["religion"])),
            "community": normalise_community(cell(ws, row, cols["community"])),
            "nationality": str_clean(cell(ws, row, cols["nationality"])),
            "place_of_birth": str_clean(cell(ws, row, cols["place_of_birth"])),
            "town_country_of_origin": str_clean(cell(ws, row, cols["town_country_of_origin"])),
            "future_settlement_plans": str_clean(cell(ws, row, cols["future_settlement"])),
            "photo_url": None,
            "photo_status": "approved",
            "status": status,
            "show_on_register": show_reg,
            "membership_expires_at": mem_exp.isoformat() if mem_exp else None,
        }

        home_line1: str | None = None
        if "home_address" in cols:
            home_line1 = str_clean(cell(ws, row, cols["home_address"]))

        private_row = {
            "surname": surname,
            "date_of_birth": dob.isoformat(),
            "email": email,
            "mobile_phone": mobile,
            "home_address_line1": home_line1,
            "father_name": str_clean(cell(ws, row, cols["father"])),
            "mother_name": str_clean(cell(ws, row, cols["mother"])),
            "id_document_url": None,
        }

        # Age sanity (optional)
        age_cell = cell(ws, row, cols["age_check"])
        if age_cell is not None and str_clean(age_cell):
            try:
                claimed = int(float(age_cell))
                years = (today - dob).days // 365
                if abs(claimed - years) > 1:
                    log.warning("Age mismatch sheet=%s computed=%s ref=%s", claimed, years, ref)
            except (TypeError, ValueError):
                pass

        display_name = f"{first} {surname} ({ref})"
        report.missing_photo_names.append(display_name)

        if args.dry_run:
            print(f"[dry-run] Would import {display_name} status={status} show_on_register={show_reg}")
            if gender == "Male":
                report.imported_male += 1
            else:
                report.imported_female += 1
            if is_expired:
                report.membership_expired += 1
            else:
                report.membership_active += 1
            continue

        assert supabase is not None
        pw = random_password_16()
        user_id: str | None = None
        try:
            uresp = supabase.auth.admin.create_user(
                {
                    "email": email,
                    "password": pw,
                    "email_confirm": True,
                }
            )
            user_id = uresp.user.id
        except Exception as e:
            msg = str(e)
            report.skipped_missing += 1
            report.errors.append(f"{ref}: auth create_user failed: {msg}")
            log.exception("create_user %s", ref)
            continue

        profile_id: str | None = None
        try:
            profile_row["auth_user_id"] = user_id
            ins = supabase.table("profiles").insert(profile_row).select("id").execute()
            if not ins.data:
                raise RuntimeError("profiles insert returned no data")
            profile_id = ins.data[0]["id"]
            priv = {**private_row, "profile_id": profile_id}
            supabase.table("member_private").insert(priv).execute()
        except Exception as e:
            report.errors.append(f"{ref}: DB insert failed: {e!r}")
            log.exception("insert profile/private %s", ref)
            try:
                if user_id:
                    supabase.auth.admin.delete_user(user_id)
            except Exception:
                log.warning("Could not delete auth user after rollback for %s", ref)
            report.skipped_missing += 1
            continue

        if gender == "Male":
            report.imported_male += 1
        else:
            report.imported_female += 1
        if is_expired:
            report.membership_expired += 1
        else:
            report.membership_active += 1

        # Recovery link (always generated for imported users)
        try:
            link_resp = supabase.auth.admin.generate_link(
                {"type": "recovery", "email": email, "options": {"redirect_to": redirect_to}}
            )
            action_link = link_resp.properties.action_link
        except Exception as e:
            action_link = f"(generate_link failed: {e!r})"
            report.errors.append(f"{ref}: generate_link failed: {e!r}")

        report.reset_links.append((ref, display_name, email, action_link))
        log.info("Imported %s", display_name)

        if args.send_welcome_emails and args.resend_api_key and action_link.startswith("http"):
            subj = "Welcome to the new Vanik Matrimonial Register"
            body = welcome_email_html(first, ref, site_url, action_link)
            mid, err = send_resend_email(args.resend_api_key, email, subj, body)
            if err:
                report.errors.append(f"{ref}: Resend failed: {err}")
            else:
                try:
                    supabase.table("email_log").insert(
                        {
                            "recipient_email": email,
                            "recipient_profile_id": profile_id,
                            "email_type": "migration_welcome",
                            "subject": subj,
                            "resend_message_id": mid,
                            "status": "sent" if mid else "failed",
                        }
                    ).execute()
                except Exception as le:
                    report.errors.append(f"{ref}: email_log insert: {le!r}")


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Migrate Matrimonial_Register.xlsx to Supabase.")
    p.add_argument(
        "--excel",
        default="Matrimonial_Register.xlsx",
        help="Path to workbook (default: ./Matrimonial_Register.xlsx)",
    )
    p.add_argument("--dry-run", action="store_true", help="Parse and report; no writes (reads DB for duplicates).")
    p.add_argument("--ref", dest="ref_filter", default=None, help='Only process this reference, e.g. "M 2516".')
    p.add_argument(
        "--send-welcome-emails",
        action="store_true",
        help="After each successful import, send welcome email via Resend (needs RESEND_API_KEY).",
    )
    p.add_argument(
        "--inspect-headers",
        action="store_true",
        help="Print row-1 headers for Boys/Girls and exit.",
    )
    p.add_argument(
        "--no-db-check",
        action="store_true",
        help="With --dry-run, do not query Supabase for existing reference numbers.",
    )
    p.add_argument(
        "--public-site-url",
        default=None,
        help="Override PUBLIC_SITE_URL for reset redirect and email links.",
    )
    p.add_argument(
        "--girls-mobile-col",
        default="AC",
        metavar="COL",
        help="Excel column on Girls sheet for mobile (default AC; verify with --inspect-headers).",
    )
    return p.parse_args(list(argv) if argv is not None else None)


def main(argv: Iterable[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    log = logging.getLogger("migrate")

    if load_dotenv:
        load_dotenv()

    import os

    args.public_site_url = args.public_site_url or os.environ.get("PUBLIC_SITE_URL")
    args.resend_api_key = os.environ.get("RESEND_API_KEY")

    excel_path = args.excel
    if not os.path.isfile(excel_path):
        log.error("Excel file not found: %s", excel_path)
        return 1

    if args.inspect_headers:
        inspect_headers(excel_path)
        return 0

    url = os.environ.get("SUPABASE_URL") or os.environ.get("VITE_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    supabase: Client | None = None
    if not args.dry_run or not args.no_db_check:
        if not url or not key:
            if args.dry_run:
                log.warning(
                    "Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY - dry-run cannot check duplicates by ref."
                )
            else:
                log.error("Set SUPABASE_URL (or VITE_SUPABASE_URL) and SUPABASE_SERVICE_ROLE_KEY in .env")
                return 1
        if url and key:
            supabase = create_client(url, key)

    if not args.dry_run and supabase is None:
        log.error("Supabase client not configured.")
        return 1

    if args.send_welcome_emails and not args.resend_api_key:
        log.error("RESEND_API_KEY required for --send-welcome-emails")
        return 1

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    report = MigrationReport()

    try:
        boys = sheet_by_name(wb, "Boys", "Boy")
        process_sheet(boys, BOYS_COLS, "Male", args, supabase, report, log)
    except KeyError as e:
        log.error("%s", e)
        return 1

    try:
        girls = sheet_by_name(wb, "Girls", "Girl")
        process_sheet(girls, girls_columns(args.girls_mobile_col), "Female", args, supabase, report, log)
    except KeyError as e:
        log.error("%s", e)
        return 1

    report.print_summary()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
